from Tkinter import *
import tkFileDialog

from openpyxl import *

import xlrd
import csv
import os



class MainWindow:

    def __init__(self, master):

        self.root = master                  # main GUI context
        self.root.title("stimuligen")       # title of window
        self.root.geometry("600x400")       # size of GUI window
        self.main_frame = Frame(root)       # main frame into which all the Gui components will be placed
        self.main_frame.pack()              # pack() basically sets up/inserts the element (turns it on)

        self.datasource_template_file = None
        self.datasource_template_book = None
        self.datasource_template_sheet = None

        self.eyetracking_order_file = None
        self.eyetracking_orders = []

        self.pair_carrier_orders_file = None
        self.pair_carrier_orders_book = None
        self.pair_carrier_orders_sheet = None

        self.load_datasource_button = Button(self.main_frame,
                                             text="Load Datasource",
                                             command=self.load_datasource)

        self.load_eyetracking_order_button = Button(self.main_frame,
                                                    text="Load Eyetracking Order",
                                                    command=self.load_eyetracking_order)

        self.clear_button = Button(self.main_frame,
                                   text="Clear",
                                   command=self.clear)

        self.load_pair_carrier_orders_button = Button(self.main_frame,
                                                      text="Load Pair Carrier Orders",
                                                      command=self.load_pair_carrier_orders)

        self.generate_stimuli_button = Button(self.main_frame,
                                              text='Generate Stimuli',
                                              command=self.generate_stimuli)


        self.video_loaded_label = Label(self.main_frame, text="video loaded", fg="blue")
        self.timestamps_loaded_label = Label(self.main_frame, text="timestamps loaded", fg="red")

        self.load_datasource_button.grid(row=1, column=1)
        self.load_eyetracking_order_button.grid(row=1, column=3)
        self.load_pair_carrier_orders_button.grid(row=2, column=2)
        self.generate_stimuli_button.grid(row=2, column=3)
        self.clear_button.grid(row=3, column=2)

    def load_datasource(self):

        self.datasource_template_file = tkFileDialog.askopenfilename()

        self.datasource_template_book = load_workbook(self.datasource_template_file)

        self.datasource_template_sheet = self.datasource_template_book.active


    def load_eyetracking_order(self):

        self.eyetracking_order_file = tkFileDialog.askopenfilename()

        with open(self.eyetracking_order_file, "rU") as file:
            csvreader = csv.reader(file, delimiter='\t')
            csvreader.next()
            for row in csvreader:
                self.eyetracking_orders.append([row[0],      # SubID
                                                row[1],      # Month
                                                int(row[2]), # Order
                                                row[3],      # carrier_order_half
                                                row[4],      # Past
                                                row[5]])     # Visit

            print self.eyetracking_orders

    def load_pair_carrier_orders(self):

        self.pair_carrier_orders_file = tkFileDialog.askopenfilename()

        self.pair_carrier_orders_book = load_workbook(self.pair_carrier_orders_file)
        self.pair_carrier_orders_sheet = self.pair_carrier_orders_book.active

    def clear(self):

        self.datasource_template_file = None
        self.eyetracking_order_file = None
        self.pair_carrier_orders_file = None

        self.eyetracking_orders = []

        self.datasource_template_book = None
        self.pair_carrier_orders_book = None

        self.datasource_template_sheet = None
        self.pair_carrier_orders_sheet = None

        if self.video_loaded_label:
            self.video_loaded_label.grid_remove()
        if self.timestamps_loaded_label:
            self.timestamps_loaded_label.grid_remove()

    def generate_stimuli(self):
        """

        This is a really stupid and explicit implementation, written just to get
        it to working condition. Everything is basically hard-coded.
        Too lazy to think about how to condense it into loops (at least for now)

        """
        header = [u'number',
                  u'word',
                  u'kind',
                  u'carrier',
                  u'duplicate_image_filename',
                  u'',
                  u'order',
                  u'pair',
                  u'pair_words',
                  u'pair_kind',
                  u'carrier']

        start8_z = 2
        start10_z = 10
        start12_z = 18
        start14_z = 26
        start16_z = 34
        start18_z = 42

        start8_y = 50
        start10_y = 58
        start12_y = 66
        start14_y = 74
        start16_y = 82
        start18_y = 90

        start8_z_uniq = 98
        start10_z_uniq = 106
        start12_z_uniq = 114
        start14_z_uniq = 122
        start16_z_uniq = 130
        start18_z_uniq = 138

        start8_y_uniq = 146
        start10_y_uniq = 154
        start12_y_uniq = 162
        start14_y_uniq = 170
        start16_y_uniq = 178
        start18_y_uniq = 186

        for entry in self.eyetracking_orders:
            if entry[4] == "past":
                continue
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(header)           # write header
                ws['H2'] = 'A'              # write pairs
                ws['H3'] = 'B'
                ws['H4'] = 'C'
                ws['H5'] = 'D'
                ws['H6'] = 'E'
                ws['H7'] = 'F'
                ws['H8'] = 'G'
                ws['H9'] = 'H'

                # write number p1-p4
                ws['A2'] = 'p1'
                ws['A3'] = 'p2'
                ws['A4'] = 'p3'
                ws['A5'] = 'p4'

                # write number 1-16
                for i in range(1, 17):
                    ws['A{}'.format(6+(i-1))] = i

                # write "practice" in C2-C5
                for j in range(4):
                    ws['C{}'.format(2+j)] = "practice"

                # write practice carriers
                ws['D2'] = "can"; ws['D3'] = "where"
                ws['D4'] = "do"; ws['D5'] = "look"

                # write "generic" in C6-C13
                for k in range(8):
                    ws['C{}'.format(6+k)] = 'generic'

                # write "generic" in J2-J5
                for l in range(4):
                    ws['J{}'.format(2+l)] = 'generic'

                # write the unique kinds (C14-C21)

                ws['C14'] = "unique_video"
                ws['C15'] = "unique_video"
                ws['C16'] = "unique_video"
                ws['C17'] = "unique_video"
                ws['C18'] = "unique_audio"
                ws['C19'] = "unique_audio"
                ws['C20'] = "unique_audio"
                ws['C21'] = "unique_audio"

                # write the unique pair_kinds (J6-J9)

                ws['J6'] = "unique_video"
                ws['J7'] = "unique_video"
                ws['J8'] = "unique_audio"
                ws['J9'] = "unique_audio"

                # write the stuff at the bottom

                ws['A27'] = "stim details"
                ws['A28'] = "month"; ws['B28'] = "word_type"; ws['C28'] = "need_audio"
                ws['D28'] = "need_image"; ws['E28'] = "word"; ws['F28'] = "count"; ws['G28'] = "find images"

                ws['A29'] = 6; ws['B29'] = "video"
                ws['A30'] = 6; ws['B30'] = "video"
                ws['A31'] = 7; ws['B31'] = "video"
                ws['A32'] = 7; ws['B32'] = "video"
                ws['A33'] = 6; ws['B33'] = "audio"
                ws['A34'] = 6; ws['B34'] = "audio"
                ws['A35'] = 7; ws['B35'] = "audio"
                ws['A36'] = 7; ws['B36'] = "audio"

                for i in range(2, 22):
                    ws['E{}'.format(i)] = "NA"

                if entry[1] == '08':
                    if entry[3] == 'Z':

                        # generic Z words
                        for l in range(8):
                            ws['B{}'.format(6+l)] = self.pair_carrier_orders_sheet['A{}'.format(start8_z+l)].value

                        # generic Z carriers
                        for l in range(8):
                            ws['D{}'.format(6+l)] = self.pair_carrier_orders_sheet['G{}'.format(start8_z+l)].value

                        # unique Z carrier
                        for l in range(8):
                            ws['D{}'.format(14+l)] = self.pair_carrier_orders_sheet['G{}'.format(start8_z_uniq+l)].value

                        # generic Z pair_words
                        ws['I2'] = self.pair_carrier_orders_sheet['C{}'.format(start8_z)].value
                        ws['I3'] = self.pair_carrier_orders_sheet['C{}'.format(start8_z+2)].value
                        ws['I4'] = self.pair_carrier_orders_sheet['C{}'.format(start8_z+4)].value
                        ws['I5'] = self.pair_carrier_orders_sheet['C{}'.format(start8_z+6)].value

                        # generic Z pair carriers
                        ws['K2'] = self.pair_carrier_orders_sheet['G{}'.format(start8_z)].value
                        ws['K3'] = self.pair_carrier_orders_sheet['G{}'.format(start8_z+2)].value
                        ws['K4'] = self.pair_carrier_orders_sheet['G{}'.format(start8_z+4)].value
                        ws['K5'] = self.pair_carrier_orders_sheet['G{}'.format(start8_z+6)].value

                        # unique Z pair carriers
                        ws['K6'] = self.pair_carrier_orders_sheet['G{}'.format(start8_z_uniq)].value
                        ws['K7'] = self.pair_carrier_orders_sheet['G{}'.format(start8_z_uniq+2)].value
                        ws['K8'] = self.pair_carrier_orders_sheet['G{}'.format(start8_z_uniq+4)].value
                        ws['K9'] = self.pair_carrier_orders_sheet['G{}'.format(start8_z_uniq+6)].value

                    else:

                        # generic Y words
                        for l in range(8):
                            ws['B{}'.format(6+l)] = self.pair_carrier_orders_sheet['A{}'.format(start8_y+l)].value

                        # generic Y carriers
                        for l in range(8):
                            ws['D{}'.format(6+l)] = self.pair_carrier_orders_sheet['G{}'.format(start8_y+l)].value

                        # unique Y carriers
                        for l in range(8):
                            ws['D{}'.format(14+l)] = self.pair_carrier_orders_sheet['G{}'.format(start8_y_uniq+l)].value

                        # generic Y pair_words
                        ws['I2'] = self.pair_carrier_orders_sheet['C{}'.format(start8_y)].value
                        ws['I3'] = self.pair_carrier_orders_sheet['C{}'.format(start8_y+2)].value
                        ws['I4'] = self.pair_carrier_orders_sheet['C{}'.format(start8_y+4)].value
                        ws['I5'] = self.pair_carrier_orders_sheet['C{}'.format(start8_y+6)].value

                        # generic Y pair carriers
                        ws['K2'] = self.pair_carrier_orders_sheet['G{}'.format(start8_y)].value
                        ws['K3'] = self.pair_carrier_orders_sheet['G{}'.format(start8_y+2)].value
                        ws['K4'] = self.pair_carrier_orders_sheet['G{}'.format(start8_y+4)].value
                        ws['K5'] = self.pair_carrier_orders_sheet['G{}'.format(start8_y+6)].value

                        # unique Y pair carriers
                        ws['K6'] = self.pair_carrier_orders_sheet['G{}'.format(start8_y_uniq)].value
                        ws['K7'] = self.pair_carrier_orders_sheet['G{}'.format(start8_y_uniq+2)].value
                        ws['K8'] = self.pair_carrier_orders_sheet['G{}'.format(start8_y_uniq+4)].value
                        ws['K9'] = self.pair_carrier_orders_sheet['G{}'.format(start8_y_uniq+6)].value

                if entry[1] == '10':
                    if entry[3] == 'Z':

                        for l in range(8):
                            ws['B{}'.format(6+l)] = self.pair_carrier_orders_sheet['A{}'.format(start10_z+l)].value

                        for l in range(8):
                            ws['D{}'.format(6+l)] = self.pair_carrier_orders_sheet['G{}'.format(start10_z+l)].value

                        for l in range(8):
                            ws['D{}'.format(14+l)] = self.pair_carrier_orders_sheet['G{}'.format(start10_z_uniq+l)].value

                        ws['I2'] = self.pair_carrier_orders_sheet['C{}'.format(start10_z)].value
                        ws['I3'] = self.pair_carrier_orders_sheet['C{}'.format(start10_z+2)].value
                        ws['I4'] = self.pair_carrier_orders_sheet['C{}'.format(start10_z+4)].value
                        ws['I5'] = self.pair_carrier_orders_sheet['C{}'.format(start10_z+6)].value

                        ws['K2'] = self.pair_carrier_orders_sheet['G{}'.format(start10_z)].value
                        ws['K3'] = self.pair_carrier_orders_sheet['G{}'.format(start10_z+2)].value
                        ws['K4'] = self.pair_carrier_orders_sheet['G{}'.format(start10_z+4)].value
                        ws['K5'] = self.pair_carrier_orders_sheet['G{}'.format(start10_z+6)].value

                        ws['K6'] = self.pair_carrier_orders_sheet['G{}'.format(start10_z_uniq)].value
                        ws['K7'] = self.pair_carrier_orders_sheet['G{}'.format(start10_z_uniq+2)].value
                        ws['K8'] = self.pair_carrier_orders_sheet['G{}'.format(start10_z_uniq+4)].value
                        ws['K9'] = self.pair_carrier_orders_sheet['G{}'.format(start10_z_uniq+6)].value
                    else:

                        for l in range(8):
                            ws['B{}'.format(6+l)] = self.pair_carrier_orders_sheet['A{}'.format(start10_y+l)].value

                        for l in range(8):
                            ws['D{}'.format(6+l)] = self.pair_carrier_orders_sheet['G{}'.format(start10_y+l)].value

                        for l in range(8):
                            ws['D{}'.format(14+l)] = self.pair_carrier_orders_sheet['G{}'.format(start10_y_uniq+l)].value

                        ws['I2'] = self.pair_carrier_orders_sheet['C{}'.format(start10_y)].value
                        ws['I3'] = self.pair_carrier_orders_sheet['C{}'.format(start10_y+2)].value
                        ws['I4'] = self.pair_carrier_orders_sheet['C{}'.format(start10_y+4)].value
                        ws['I5'] = self.pair_carrier_orders_sheet['C{}'.format(start10_y+6)].value

                        ws['K2'] = self.pair_carrier_orders_sheet['G{}'.format(start10_y)].value
                        ws['K3'] = self.pair_carrier_orders_sheet['G{}'.format(start10_y+2)].value
                        ws['K4'] = self.pair_carrier_orders_sheet['G{}'.format(start10_y+4)].value
                        ws['K5'] = self.pair_carrier_orders_sheet['G{}'.format(start10_y+6)].value

                        ws['K6'] = self.pair_carrier_orders_sheet['G{}'.format(start10_y_uniq)].value
                        ws['K7'] = self.pair_carrier_orders_sheet['G{}'.format(start10_y_uniq+2)].value
                        ws['K8'] = self.pair_carrier_orders_sheet['G{}'.format(start10_y_uniq+4)].value
                        ws['K9'] = self.pair_carrier_orders_sheet['G{}'.format(start10_y_uniq+6)].value

                if entry[1] == '12':
                    if entry[3] == 'Z':

                        for l in range(8):
                            ws['B{}'.format(6+l)] = self.pair_carrier_orders_sheet['A{}'.format(start12_z+l)].value

                        for l in range(8):
                            ws['D{}'.format(6+l)] = self.pair_carrier_orders_sheet['G{}'.format(start12_z+l)].value

                        for l in range(8):
                            ws['D{}'.format(14+l)] = self.pair_carrier_orders_sheet['G{}'.format(start12_z_uniq+l)].value

                        ws['I2'] = self.pair_carrier_orders_sheet['C{}'.format(start12_z)].value
                        ws['I3'] = self.pair_carrier_orders_sheet['C{}'.format(start12_z+2)].value
                        ws['I4'] = self.pair_carrier_orders_sheet['C{}'.format(start12_z+4)].value
                        ws['I5'] = self.pair_carrier_orders_sheet['C{}'.format(start12_z+6)].value

                        ws['K2'] = self.pair_carrier_orders_sheet['G{}'.format(start12_z)].value
                        ws['K3'] = self.pair_carrier_orders_sheet['G{}'.format(start12_z+2)].value
                        ws['K4'] = self.pair_carrier_orders_sheet['G{}'.format(start12_z+4)].value
                        ws['K5'] = self.pair_carrier_orders_sheet['G{}'.format(start12_z+6)].value

                        ws['K6'] = self.pair_carrier_orders_sheet['G{}'.format(start12_z_uniq)].value
                        ws['K7'] = self.pair_carrier_orders_sheet['G{}'.format(start12_z_uniq+2)].value
                        ws['K8'] = self.pair_carrier_orders_sheet['G{}'.format(start12_z_uniq+4)].value
                        ws['K9'] = self.pair_carrier_orders_sheet['G{}'.format(start12_z_uniq+6)].value
                    else:

                        for l in range(8):
                            ws['B{}'.format(6+l)] = self.pair_carrier_orders_sheet['A{}'.format(start12_y+l)].value

                        for l in range(8):
                            ws['D{}'.format(6+l)] = self.pair_carrier_orders_sheet['G{}'.format(start12_y+l)].value

                        for l in range(8):
                            ws['D{}'.format(14+l)] = self.pair_carrier_orders_sheet['G{}'.format(start12_y_uniq+l)].value

                        ws['I2'] = self.pair_carrier_orders_sheet['C{}'.format(start12_y)].value
                        ws['I3'] = self.pair_carrier_orders_sheet['C{}'.format(start12_y+2)].value
                        ws['I4'] = self.pair_carrier_orders_sheet['C{}'.format(start12_y+4)].value
                        ws['I5'] = self.pair_carrier_orders_sheet['C{}'.format(start12_y+6)].value

                        ws['K2'] = self.pair_carrier_orders_sheet['G{}'.format(start12_y)].value
                        ws['K3'] = self.pair_carrier_orders_sheet['G{}'.format(start12_y+2)].value
                        ws['K4'] = self.pair_carrier_orders_sheet['G{}'.format(start12_y+4)].value
                        ws['K5'] = self.pair_carrier_orders_sheet['G{}'.format(start12_y+6)].value

                        ws['K6'] = self.pair_carrier_orders_sheet['G{}'.format(start12_y_uniq)].value
                        ws['K7'] = self.pair_carrier_orders_sheet['G{}'.format(start12_y_uniq+2)].value
                        ws['K8'] = self.pair_carrier_orders_sheet['G{}'.format(start12_y_uniq+4)].value
                        ws['K9'] = self.pair_carrier_orders_sheet['G{}'.format(start12_y_uniq+6)].value

                if entry[1] == '14':
                    if entry[3] == 'Z':

                        for l in range(8):
                            ws['B{}'.format(6+l)] = self.pair_carrier_orders_sheet['A{}'.format(start14_z+l)].value

                        for l in range(8):
                            ws['D{}'.format(6+l)] = self.pair_carrier_orders_sheet['G{}'.format(start14_z+l)].value

                        for l in range(8):
                            ws['D{}'.format(14+l)] = self.pair_carrier_orders_sheet['G{}'.format(start14_z_uniq+l)].value

                        ws['I2'] = self.pair_carrier_orders_sheet['C{}'.format(start14_z)].value
                        ws['I3'] = self.pair_carrier_orders_sheet['C{}'.format(start14_z+2)].value
                        ws['I4'] = self.pair_carrier_orders_sheet['C{}'.format(start14_z+4)].value
                        ws['I5'] = self.pair_carrier_orders_sheet['C{}'.format(start14_z+6)].value

                        ws['K2'] = self.pair_carrier_orders_sheet['G{}'.format(start14_z)].value
                        ws['K3'] = self.pair_carrier_orders_sheet['G{}'.format(start14_z+2)].value
                        ws['K4'] = self.pair_carrier_orders_sheet['G{}'.format(start14_z+4)].value
                        ws['K5'] = self.pair_carrier_orders_sheet['G{}'.format(start14_z+6)].value

                        ws['K6'] = self.pair_carrier_orders_sheet['G{}'.format(start14_z_uniq)].value
                        ws['K7'] = self.pair_carrier_orders_sheet['G{}'.format(start14_z_uniq+2)].value
                        ws['K8'] = self.pair_carrier_orders_sheet['G{}'.format(start14_z_uniq+4)].value
                        ws['K9'] = self.pair_carrier_orders_sheet['G{}'.format(start14_z_uniq+6)].value
                    else:

                        for l in range(8):
                            ws['B{}'.format(6+l)] = self.pair_carrier_orders_sheet['A{}'.format(start14_y+l)].value

                        for l in range(8):
                            ws['D{}'.format(6+l)] = self.pair_carrier_orders_sheet['G{}'.format(start14_y+l)].value

                        for l in range(8):
                            ws['D{}'.format(14+l)] = self.pair_carrier_orders_sheet['G{}'.format(start14_y_uniq+l)].value

                        ws['I2'] = self.pair_carrier_orders_sheet['C{}'.format(start14_y)].value
                        ws['I3'] = self.pair_carrier_orders_sheet['C{}'.format(start14_y+2)].value
                        ws['I4'] = self.pair_carrier_orders_sheet['C{}'.format(start14_y+4)].value
                        ws['I5'] = self.pair_carrier_orders_sheet['C{}'.format(start14_y+6)].value

                        ws['K2'] = self.pair_carrier_orders_sheet['G{}'.format(start14_y)].value
                        ws['K3'] = self.pair_carrier_orders_sheet['G{}'.format(start14_y+2)].value
                        ws['K4'] = self.pair_carrier_orders_sheet['G{}'.format(start14_y+4)].value
                        ws['K5'] = self.pair_carrier_orders_sheet['G{}'.format(start14_y+6)].value

                        ws['K6'] = self.pair_carrier_orders_sheet['G{}'.format(start14_y_uniq)].value
                        ws['K7'] = self.pair_carrier_orders_sheet['G{}'.format(start14_y_uniq+2)].value
                        ws['K8'] = self.pair_carrier_orders_sheet['G{}'.format(start14_y_uniq+4)].value
                        ws['K9'] = self.pair_carrier_orders_sheet['G{}'.format(start14_y_uniq+6)].value

                if entry[1] == '16':
                    if entry[3] == 'Z':

                        for l in range(8):
                            ws['B{}'.format(6+l)] = self.pair_carrier_orders_sheet['A{}'.format(start16_z+l)].value

                        for l in range(8):
                            ws['D{}'.format(6+l)] = self.pair_carrier_orders_sheet['G{}'.format(start16_z+l)].value

                        for l in range(8):
                            ws['D{}'.format(14+l)] = self.pair_carrier_orders_sheet['G{}'.format(start16_z_uniq+l)].value

                        ws['I2'] = self.pair_carrier_orders_sheet['C{}'.format(start16_z)].value
                        ws['I3'] = self.pair_carrier_orders_sheet['C{}'.format(start16_z+2)].value
                        ws['I4'] = self.pair_carrier_orders_sheet['C{}'.format(start16_z+4)].value
                        ws['I5'] = self.pair_carrier_orders_sheet['C{}'.format(start16_z+6)].value

                        ws['K2'] = self.pair_carrier_orders_sheet['G{}'.format(start16_z)].value
                        ws['K3'] = self.pair_carrier_orders_sheet['G{}'.format(start16_z+2)].value
                        ws['K4'] = self.pair_carrier_orders_sheet['G{}'.format(start16_z+4)].value
                        ws['K5'] = self.pair_carrier_orders_sheet['G{}'.format(start16_z+6)].value

                        ws['K6'] = self.pair_carrier_orders_sheet['G{}'.format(start16_z_uniq)].value
                        ws['K7'] = self.pair_carrier_orders_sheet['G{}'.format(start16_z_uniq+2)].value
                        ws['K8'] = self.pair_carrier_orders_sheet['G{}'.format(start16_z_uniq+4)].value
                        ws['K9'] = self.pair_carrier_orders_sheet['G{}'.format(start16_z_uniq+6)].value

                    else:

                        for l in range(8):
                            ws['B{}'.format(6+l)] = self.pair_carrier_orders_sheet['A{}'.format(start16_y+l)].value

                        for l in range(8):
                            ws['D{}'.format(6+l)] = self.pair_carrier_orders_sheet['G{}'.format(start16_y+l)].value

                        for l in range(8):
                            ws['D{}'.format(14+l)] = self.pair_carrier_orders_sheet['G{}'.format(start16_y_uniq+l)].value

                        ws['I2'] = self.pair_carrier_orders_sheet['C{}'.format(start16_y)].value
                        ws['I3'] = self.pair_carrier_orders_sheet['C{}'.format(start16_y+2)].value
                        ws['I4'] = self.pair_carrier_orders_sheet['C{}'.format(start16_y+4)].value
                        ws['I5'] = self.pair_carrier_orders_sheet['C{}'.format(start16_y+6)].value

                        ws['K2'] = self.pair_carrier_orders_sheet['G{}'.format(start16_y)].value
                        ws['K3'] = self.pair_carrier_orders_sheet['G{}'.format(start16_y+2)].value
                        ws['K4'] = self.pair_carrier_orders_sheet['G{}'.format(start16_y+4)].value
                        ws['K5'] = self.pair_carrier_orders_sheet['G{}'.format(start16_y+6)].value

                        ws['K6'] = self.pair_carrier_orders_sheet['G{}'.format(start16_y_uniq)].value
                        ws['K7'] = self.pair_carrier_orders_sheet['G{}'.format(start16_y_uniq+2)].value
                        ws['K8'] = self.pair_carrier_orders_sheet['G{}'.format(start16_y_uniq+4)].value
                        ws['K9'] = self.pair_carrier_orders_sheet['G{}'.format(start16_y_uniq+6)].value

                if entry[1] == '18':
                    if entry[3] == 'Z':

                        for l in range(8):
                            ws['B{}'.format(6+l)] = self.pair_carrier_orders_sheet['A{}'.format(start18_z+l)].value

                        for l in range(8):
                            ws['D{}'.format(6+l)] = self.pair_carrier_orders_sheet['G{}'.format(start18_z+l)].value

                        for l in range(8):
                            ws['D{}'.format(14+l)] = self.pair_carrier_orders_sheet['G{}'.format(start18_z_uniq+l)].value

                        ws['I2'] = self.pair_carrier_orders_sheet['C{}'.format(start18_z)].value
                        ws['I3'] = self.pair_carrier_orders_sheet['C{}'.format(start18_z+2)].value
                        ws['I4'] = self.pair_carrier_orders_sheet['C{}'.format(start18_z+4)].value
                        ws['I5'] = self.pair_carrier_orders_sheet['C{}'.format(start18_z+6)].value

                        ws['K2'] = self.pair_carrier_orders_sheet['G{}'.format(start18_z)].value
                        ws['K3'] = self.pair_carrier_orders_sheet['G{}'.format(start18_z+2)].value
                        ws['K4'] = self.pair_carrier_orders_sheet['G{}'.format(start18_z+4)].value
                        ws['K5'] = self.pair_carrier_orders_sheet['G{}'.format(start18_z+6)].value

                        ws['K6'] = self.pair_carrier_orders_sheet['G{}'.format(start18_z_uniq)].value
                        ws['K7'] = self.pair_carrier_orders_sheet['G{}'.format(start18_z_uniq+2)].value
                        ws['K8'] = self.pair_carrier_orders_sheet['G{}'.format(start18_z_uniq+4)].value
                        ws['K9'] = self.pair_carrier_orders_sheet['G{}'.format(start18_z_uniq+6)].value
                    else:

                        for l in range(8):
                            ws['B{}'.format(6+l)] = self.pair_carrier_orders_sheet['A{}'.format(start18_y+l)].value

                        for l in range(8):
                            ws['D{}'.format(6+l)] = self.pair_carrier_orders_sheet['G{}'.format(start18_y+l)].value

                        for l in range(8):
                            ws['D{}'.format(14+l)] = self.pair_carrier_orders_sheet['G{}'.format(start18_y_uniq+l)].value

                        ws['I2'] = self.pair_carrier_orders_sheet['C{}'.format(start18_y)].value
                        ws['I3'] = self.pair_carrier_orders_sheet['C{}'.format(start18_y+2)].value
                        ws['I4'] = self.pair_carrier_orders_sheet['C{}'.format(start18_y+4)].value
                        ws['I5'] = self.pair_carrier_orders_sheet['C{}'.format(start18_y+6)].value

                        ws['K2'] = self.pair_carrier_orders_sheet['G{}'.format(start18_y)].value
                        ws['K3'] = self.pair_carrier_orders_sheet['G{}'.format(start18_y+2)].value
                        ws['K4'] = self.pair_carrier_orders_sheet['G{}'.format(start18_y+4)].value
                        ws['K5'] = self.pair_carrier_orders_sheet['G{}'.format(start18_y+6)].value

                        ws['K6'] = self.pair_carrier_orders_sheet['G{}'.format(start18_y_uniq)].value
                        ws['K7'] = self.pair_carrier_orders_sheet['G{}'.format(start18_y_uniq+2)].value
                        ws['K8'] = self.pair_carrier_orders_sheet['G{}'.format(start18_y_uniq+4)].value
                        ws['K9'] = self.pair_carrier_orders_sheet['G{}'.format(start18_y_uniq+6)].value

                ws['G2'] = entry[2]         # write order
                wb.save("output/{}_stimuli.xlsx".format(entry[5]))  # export xlsx file

if __name__ == "__main__":

    root = Tk()
    MainWindow(root)
    root.mainloop()
